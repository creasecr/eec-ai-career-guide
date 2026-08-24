import json
import re
import sys
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


def clean(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value).strip()


def normalize_level(value):
    v = clean(value).lower()
    return {"basic": "Beginner", "beginner": "Beginner", "intermediate": "Intermediate", "advanced": "Advanced"}.get(v, clean(value) or "Unspecified")


def infer_delivery(location, event_tags):
    text = f"{clean(location)} {clean(event_tags)}".lower()
    if any(x in text for x in ("online", "webinar", "zoom", "adobe connect", "teams")):
        return "Virtual"
    return "In-Person"


def infer_topics(title, category):
    text = f"{clean(title)} {clean(category)}".lower()
    rules = [
        ("EV Charging", ("electric vehicle", "ev charging", "bi-directional", "bidirectional")),
        ("Solar & Storage", ("solar", "battery", "storage", "nem")),
        ("Electrification & Decarbonization", ("electrification", "all-electric", "decarbon", "heat pump", "emissions")),
        ("Agriculture, Pumps & Irrigation", ("agriculture", "ag customer", "irrigation", "pump test", "flowmeter", "water")),
        ("Energy Efficiency & Building Operations", ("energy efficien", "building operator", "boc ", "benchmarking", "building performance")),
        ("HVAC/R", ("hvac", "refriger", "chiller", "air balancing", "duct", "combustion", "hydronic", "ventilation")),
        ("Energy Codes & Standards", ("title 24", "calgreen", "energy code", "energypro", "cbecc", "lighting requirements")),
        ("Industrial Automation & Technology", ("automation", "controller", "robot", "fanuc", "industrial wiring", "analog signals")),
        ("Foodservice", ("foodservice", "cookline", "culinary", "induction", "pizza", "ice machine")),
        ("Lighting", ("lighting",)),
        ("Finance & Sales", ("lending", "finance", "sales")),
    ]
    topics = [name for name, words in rules if any(word in text for word in words)]
    return topics or [clean(category) or "Other"]


def audiences(sector, category):
    result = []
    s = clean(sector).lower()
    c = clean(category).lower()
    if "residential" in s and "non" not in s:
        result += ["Residential Customer", "Contractor"]
    if "non-residential" in s or "nonresidential" in s:
        result += ["Commercial Customer", "Government", "Contractor"]
    if "combination" in s:
        result += ["Residential Customer", "Commercial Customer", "Contractor"]
    if "agriculture" in c or "irrigation" in c:
        result += ["Agricultural Customer", "Commercial Customer"]
    if "codes" in c or "hvac" in c or "lighting" in c:
        result += ["Contractor", "Building Professional"]
    return list(dict.fromkeys(result)) or ["General Learner"]


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python3 convert_momentus_to_json.py <input.xlsx> [output.json]")
    source = Path(sys.argv[1])
    output = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("courses.json")

    wb = load_workbook(source, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header_index = next(i for i, row in enumerate(rows) if "Event ID" in [clean(v) for v in row])
    headers = [clean(v) for v in rows[header_index]]
    positions = {name: i for i, name in enumerate(headers) if name}

    def value(row, name):
        idx = positions.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    courses = []
    seen = set()
    for row in rows[header_index + 1:]:
        event_id = clean(value(row, "Event ID"))
        title = clean(value(row, "Description"))
        if not re.fullmatch(r"\d+", event_id) or not title:
            continue
        if event_id in seen:
            continue
        seen.add(event_id)

        category = clean(value(row, "Type")) or clean(value(row, "Category"))
        location = clean(value(row, "Location"))
        event_tags = clean(value(row, "Event Type Tags"))
        sector = clean(value(row, "Sector Category"))
        virtual_link = clean(value(row, "Virtual Link"))

        course = {
            "id": event_id,
            "title": title,
            "startDate": clean(value(row, "Start Date")),
            "startTime": clean(value(row, "Start Time")),
            "endTime": clean(value(row, "End Time")),
            "category": category,
            "topics": infer_topics(title, category),
            "audiences": audiences(sector, category),
            "experienceLevel": normalize_level(value(row, "Learning Level")),
            "deliveryType": infer_delivery(location, event_tags),
            "location": location,
            "sector": sector,
            "learningUnits": clean(value(row, "Learning Units")),
            "status": clean(value(row, "Status")),
            "instructor": clean(value(row, "Instructor Name")),
            "registrationUrl": f"https://sce.ungerboeck.com/prod/emc00/EventSearch.htm?mid=1&EvtID={event_id}",
            "virtualLink": virtual_link,
            "searchText": " ".join(filter(None, [title, category, sector, location])).lower()
        }
        courses.append(course)

    output.write_text(json.dumps(courses, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Created {output} with {len(courses)} courses.")


if __name__ == "__main__":
    main()
